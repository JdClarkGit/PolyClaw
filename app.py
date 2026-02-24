#!/usr/bin/env python3
"""
PolyClaw - AI-Powered Trading Strategy Platform
Strategy ideation, analysis, diagnosis, and iteration for Polymarket.
"""

from flask import Flask, jsonify, request, send_from_directory, send_file, Response, redirect
import requests
import json
import csv
import io
import os
import time
import threading
from datetime import datetime, timezone
from typing import List, Dict, Set, Optional
from collections import defaultdict
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

from analytics import (
    analyze_trades, generate_report, compare_wallets,
    PRICING_TIERS, get_tier_info, check_feature_access, calculate_overage
)
from ai_analysis import run_ai_analysis, get_available_providers
from strategy_engine import (
    STRATEGY_TEMPLATES, IDEATION_PROMPTS,
    analyze_wallet_strategy, generate_strategy_ideas,
    iterate_strategy, define_strategy
)
from notifications import (
    add_discord_webhook, remove_discord_webhook,
    add_telegram_config, remove_telegram_config,
    subscribe_to_wallet, unsubscribe_from_wallet,
    notify_trade, get_notification_channels
)

# Real-time tracking state
tracked_wallets: Dict[str, Dict] = {}
wallet_update_callbacks: Dict[str, List] = defaultdict(list)
from terminal_analytics import run_terminal_analysis

app = Flask(__name__, static_folder='.')

# Configuration
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///polyedge.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Initialize database
from models import db, init_db
init_db(app)

# Initialize authentication
from auth import init_auth
init_auth(app)

# Initialize payments
from payments import init_payments
init_payments(app)

DATA_API = "https://data-api.polymarket.com"
OUTPUT_DIR = "activity-exports"
BATCH_SIZE = 1000


def filter_trades_by_date(result: Dict, start_date: str = None, end_date: str = None) -> Dict:
    """Filter trades by date range. Dates should be in YYYY-MM-DD format."""
    if not result.get('trades'):
        return result
    
    filtered_trades = []
    start_ts = None
    end_ts = None
    
    if start_date:
        try:
            start_ts = datetime.strptime(start_date, '%Y-%m-%d').replace(tzinfo=timezone.utc).timestamp()
        except ValueError:
            pass
    
    if end_date:
        try:
            # End of day for end_date
            end_ts = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59, tzinfo=timezone.utc).timestamp()
        except ValueError:
            pass
    
    for trade in result['trades']:
        ts = trade.get('timestamp', 0)
        if start_ts and ts < start_ts:
            continue
        if end_ts and ts > end_ts:
            continue
        filtered_trades.append(trade)
    
    # Update result with filtered trades
    result['trades'] = filtered_trades
    result['trade_count'] = len(filtered_trades)
    
    # Update stats if present
    if filtered_trades and result.get('stats'):
        oldest = min(t.get('timestamp', 0) for t in filtered_trades)
        newest = max(t.get('timestamp', 0) for t in filtered_trades)
        result['stats']['oldest_trade'] = datetime.fromtimestamp(oldest).strftime('%Y-%m-%d') if oldest else None
        result['stats']['newest_trade'] = datetime.fromtimestamp(newest).strftime('%Y-%m-%d') if newest else None
    
    return result


def fetch_all_trades(wallet: str, callback=None) -> Dict:
    """Fetch ALL trades using backward timestamp pagination."""

    all_trades = []
    seen_keys: Set[tuple] = set()
    end_ts = None
    username = None
    batch_num = 0

    while True:
        batch_num += 1

        # Build request
        params = {"user": wallet, "limit": BATCH_SIZE}
        if end_ts:
            params["end"] = end_ts

        # Retry logic with exponential backoff
        trades = None
        max_retries = 3
        for retry in range(max_retries):
            try:
                response = requests.get(f"{DATA_API}/activity", params=params, timeout=15, proxies={"http": None, "https": None})
                response.raise_for_status()
                trades = response.json()
                break
            except Exception as e:
                if retry < max_retries - 1:
                    time.sleep(2 ** retry)  # Exponential backoff: 1s, 2s, 4s
                    continue
                else:
                    # Return empty result on failure
                    return {
                        "wallet": wallet,
                        "username": None,
                        "fetched_at": datetime.utcnow().isoformat() + "Z",
                        "trade_count": 0,
                        "trades": [],
                        "error": str(e)
                    }

        if not trades or not isinstance(trades, list) or len(trades) == 0:
            # No trades found - return quickly
            if batch_num == 1:
                return {
                    "wallet": wallet,
                    "username": None,
                    "fetched_at": datetime.utcnow().isoformat() + "Z",
                    "trade_count": 0,
                    "trades": [],
                    "message": "No trading activity found for this wallet on Polymarket"
                }
            break

        # Get username
        if not username and trades[0].get('name'):
            username = trades[0].get('name')

        # Deduplicate and add
        new_count = 0
        oldest_ts = float('inf')

        for t in trades:
            key = (t.get('transactionHash'), t.get('timestamp'), t.get('asset'))
            if key not in seen_keys:
                seen_keys.add(key)
                all_trades.append(t)
                new_count += 1

            ts = t.get('timestamp', 0)
            if ts and ts < oldest_ts:
                oldest_ts = ts

        # Check termination conditions
        if len(trades) < BATCH_SIZE or new_count == 0:
            break

        # Set next end timestamp
        if oldest_ts != float('inf'):
            end_ts = oldest_ts
        else:
            break

        # Small delay to be nice to API
        time.sleep(0.2)

        # Safety limit (prevent infinite loops)
        if batch_num > 5000:
            break

    # Sort by timestamp descending
    all_trades.sort(key=lambda t: t.get('timestamp', 0), reverse=True)

    # Calculate stats
    buys = sum(1 for t in all_trades if t.get('side') == 'BUY')
    sells = sum(1 for t in all_trades if t.get('side') == 'SELL')
    volume = sum(t.get('usdcSize', 0) or 0 for t in all_trades)

    timestamps = [t.get('timestamp', 0) for t in all_trades if t.get('timestamp')]
    oldest = datetime.fromtimestamp(min(timestamps)).strftime('%Y-%m-%d %H:%M') if timestamps else None
    newest = datetime.fromtimestamp(max(timestamps)).strftime('%Y-%m-%d %H:%M') if timestamps else None

    return {
        "wallet": wallet,
        "username": username,
        "fetched_at": datetime.now(timezone.utc).isoformat(),
        "trade_count": len(all_trades),
        "stats": {
            "buys": buys,
            "sells": sells,
            "volume_usd": volume,
            "oldest_trade": oldest,
            "newest_trade": newest,
            "batches": batch_num
        },
        "trades": all_trades
    }


def fetch_recent_trades(wallet: str, limit: int = 100) -> Dict:
    """Fetch recent trades (quick mode)."""
    params = {"user": wallet, "limit": limit}

    try:
        response = requests.get(f"{DATA_API}/activity", params=params, timeout=15, proxies={"http": None, "https": None})
        response.raise_for_status()
        trades = response.json() if isinstance(response.json(), list) else []

        username = None
        if trades and trades[0].get('name'):
            username = trades[0].get('name')

        # Handle empty wallets with clear message
        if not trades:
            return {
                "wallet": wallet,
                "username": None,
                "fetched_at": datetime.now(timezone.utc).isoformat(),
                "trade_count": 0,
                "trades": [],
                "message": "No trading activity found for this wallet on Polymarket"
            }

        return {
            "wallet": wallet,
            "username": username,
            "fetched_at": datetime.now(timezone.utc).isoformat(),
            "trade_count": len(trades),
            "trades": trades[:limit]
        }
    except Exception as e:
        return {"wallet": wallet, "error": str(e), "trades": [], "trade_count": 0}


def fetch_trades_stream(wallet: str, target_limit: int = None):
    """Generator function for SSE streaming progress during trade fetching."""
    all_trades = []
    seen_keys: Set[tuple] = set()
    end_ts = None
    username = None
    batch_num = 0
    start_time = time.time()

    while True:
        batch_num += 1

        # Build request
        params = {"user": wallet, "limit": BATCH_SIZE}
        if end_ts:
            params["end"] = end_ts

        # Send progress update
        elapsed = time.time() - start_time
        estimated_batches = max(batch_num, 10)  # Minimum estimate
        estimated_remaining = max(0, (elapsed / batch_num) * (estimated_batches - batch_num))
        
        progress_data = {
            "type": "progress",
            "batch": batch_num,
            "total_trades": len(all_trades),
            "elapsed_seconds": round(elapsed, 1),
            "estimated_remaining": round(estimated_remaining, 1)
        }
        yield f"data: {json.dumps(progress_data)}\n\n"

        # Retry logic with exponential backoff
        trades = None
        max_retries = 3
        for retry in range(max_retries):
            try:
                response = requests.get(f"{DATA_API}/activity", params=params, timeout=30, proxies={"http": None, "https": None})
                response.raise_for_status()
                trades = response.json()
                break
            except Exception as e:
                if retry < max_retries - 1:
                    time.sleep(2 ** retry)
                    continue
                else:
                    break

        if not trades or not isinstance(trades, list):
            break

        # Get username
        if not username and trades[0].get('name'):
            username = trades[0].get('name')

        # Deduplicate and add
        new_count = 0
        oldest_ts = float('inf')

        for t in trades:
            key = (t.get('transactionHash'), t.get('timestamp'), t.get('asset'))
            if key not in seen_keys:
                seen_keys.add(key)
                all_trades.append(t)
                new_count += 1

            ts = t.get('timestamp', 0)
            if ts and ts < oldest_ts:
                oldest_ts = ts

        # Check termination conditions
        if len(trades) < BATCH_SIZE or new_count == 0:
            break

        # Check if we've hit target limit
        if target_limit and len(all_trades) >= target_limit:
            all_trades = all_trades[:target_limit]
            break

        # Set next end timestamp
        if oldest_ts != float('inf'):
            end_ts = oldest_ts
        else:
            break

        time.sleep(0.2)

        # Safety limit
        if batch_num > 5000:
            break

    # Sort by timestamp descending
    all_trades.sort(key=lambda t: t.get('timestamp', 0), reverse=True)

    # Send completion event
    complete_data = {
        "type": "complete",
        "wallet": wallet,
        "username": username,
        "trade_count": len(all_trades),
        "total_batches": batch_num,
        "elapsed_seconds": round(time.time() - start_time, 1)
    }
    yield f"data: {json.dumps(complete_data)}\n\n"


@app.route('/')
def index():
    return send_from_directory('.', 'trade-viewer.html')


@app.route('/dashboard')
def dashboard():
    return redirect('/')


@app.route('/terminal')
def terminal_mode():
    return send_from_directory('.', 'terminal-mode.html')


@app.route('/leaderboard')
def leaderboard_page():
    return send_from_directory('.', 'leaderboard.html')


# Leaderboard data file
LEADERBOARD_FILE = os.path.join(os.path.dirname(__file__), 'leaderboard_data.json')

def load_leaderboard_data():
    """Load leaderboard data from file."""
    if os.path.exists(LEADERBOARD_FILE):
        try:
            with open(LEADERBOARD_FILE, 'r') as f:
                return json.load(f)
        except:
            pass
    return {"wallets": {}, "last_updated": None}

def save_leaderboard_data(data):
    """Save leaderboard data to file."""
    data["last_updated"] = datetime.now(timezone.utc).isoformat()
    with open(LEADERBOARD_FILE, 'w') as f:
        json.dump(data, f, indent=2)


@app.route('/api/leaderboard')
def get_leaderboard():
    """Get leaderboard rankings."""
    period = request.args.get('period', '30d')
    
    data = load_leaderboard_data()
    wallets_data = data.get("wallets", {})
    
    # Convert to list and add computed fields
    wallets = []
    total_volume = 0
    total_pnl = 0
    win_rates = []
    
    for addr, info in wallets_data.items():
        wallet_info = {
            "wallet": addr,
            "username": info.get("username"),
            "net_pnl": info.get("net_pnl", 0),
            "win_rate": info.get("win_rate", 0),
            "total_volume": info.get("total_volume", 0),
            "total_trades": info.get("total_trades", 0),
            "sharpe_like": info.get("sharpe_like", 0),
            "trader_type": info.get("trader_type", "unknown"),
            "last_updated": info.get("last_updated")
        }
        wallets.append(wallet_info)
        total_volume += wallet_info["total_volume"]
        total_pnl = max(total_pnl, wallet_info["net_pnl"])
        if wallet_info["win_rate"] > 0:
            win_rates.append(wallet_info["win_rate"])
    
    # Sort by P&L by default
    wallets.sort(key=lambda x: x.get("net_pnl", 0), reverse=True)
    
    return jsonify({
        "wallets": wallets,
        "stats": {
            "total_traders": len(wallets),
            "total_volume": total_volume,
            "avg_win_rate": sum(win_rates) / len(win_rates) if win_rates else 0,
            "top_pnl": total_pnl
        },
        "period": period
    })


@app.route('/api/leaderboard/submit', methods=['POST'])
def submit_to_leaderboard():
    """Submit a wallet to the leaderboard."""
    req_data = request.get_json()
    wallet = req_data.get('wallet', '').strip()
    
    if not wallet:
        return jsonify({"error": "Wallet address required"}), 400
    
    # Fetch and analyze the wallet
    try:
        result = fetch_trades_with_limit(wallet, 1000)
        trades = result.get('trades', [])
        
        if not trades:
            return jsonify({"error": "No trades found for this wallet"}), 404
        
        # Analyze the wallet
        analysis = analyze_trades(trades)
        
        # Load current leaderboard
        data = load_leaderboard_data()
        
        # Add/update wallet
        data["wallets"][wallet] = {
            "username": result.get("username"),
            "net_pnl": analysis.get("net_pnl", 0),
            "win_rate": analysis.get("win_rate", 0),
            "total_volume": analysis.get("total_volume", 0),
            "total_trades": analysis.get("total_trades", 0),
            "sharpe_like": analysis.get("sharpe_like", 0),
            "max_drawdown": analysis.get("max_drawdown", 0),
            "trader_type": analysis.get("trader_type", "unknown"),
            "last_updated": datetime.now(timezone.utc).isoformat()
        }
        
        # Save
        save_leaderboard_data(data)
        
        return jsonify({
            "success": True,
            "wallet": wallet,
            "username": result.get("username"),
            "stats": data["wallets"][wallet]
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/leaderboard/refresh/<wallet>', methods=['POST'])
def refresh_leaderboard_wallet(wallet):
    """Refresh stats for a wallet on the leaderboard."""
    data = load_leaderboard_data()
    
    if wallet not in data.get("wallets", {}):
        return jsonify({"error": "Wallet not on leaderboard"}), 404
    
    try:
        result = fetch_trades_with_limit(wallet, 1000)
        trades = result.get('trades', [])
        
        if not trades:
            return jsonify({"error": "No trades found"}), 404
        
        analysis = analyze_trades(trades)
        
        data["wallets"][wallet] = {
            "username": result.get("username"),
            "net_pnl": analysis.get("net_pnl", 0),
            "win_rate": analysis.get("win_rate", 0),
            "total_volume": analysis.get("total_volume", 0),
            "total_trades": analysis.get("total_trades", 0),
            "sharpe_like": analysis.get("sharpe_like", 0),
            "max_drawdown": analysis.get("max_drawdown", 0),
            "trader_type": analysis.get("trader_type", "unknown"),
            "last_updated": datetime.now(timezone.utc).isoformat()
        }
        
        save_leaderboard_data(data)
        
        return jsonify({"success": True, "stats": data["wallets"][wallet]})
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ============ NOTIFICATIONS API ============

@app.route('/api/notifications/channels')
def get_channels():
    """Get all configured notification channels."""
    return jsonify(get_notification_channels())


@app.route('/api/notifications/discord', methods=['POST'])
def add_discord():
    """Add a Discord webhook."""
    data = request.get_json()
    name = data.get('name', '').strip()
    webhook_url = data.get('webhook_url', '').strip()
    
    if not name or not webhook_url:
        return jsonify({"error": "Name and webhook_url required"}), 400
    
    result = add_discord_webhook(name, webhook_url)
    if result.get("error"):
        return jsonify(result), 400
    return jsonify(result)


@app.route('/api/notifications/discord/<name>', methods=['DELETE'])
def delete_discord(name):
    """Remove a Discord webhook."""
    result = remove_discord_webhook(name)
    if result.get("error"):
        return jsonify(result), 404
    return jsonify(result)


@app.route('/api/notifications/telegram', methods=['POST'])
def add_telegram():
    """Add a Telegram bot configuration."""
    data = request.get_json()
    name = data.get('name', '').strip()
    bot_token = data.get('bot_token', '').strip()
    chat_id = data.get('chat_id', '').strip()
    
    if not name or not bot_token or not chat_id:
        return jsonify({"error": "Name, bot_token, and chat_id required"}), 400
    
    result = add_telegram_config(name, bot_token, chat_id)
    if result.get("error"):
        return jsonify(result), 400
    return jsonify(result)


@app.route('/api/notifications/telegram/<name>', methods=['DELETE'])
def delete_telegram(name):
    """Remove a Telegram configuration."""
    result = remove_telegram_config(name)
    if result.get("error"):
        return jsonify(result), 404
    return jsonify(result)


@app.route('/api/notifications/subscribe', methods=['POST'])
def subscribe_wallet():
    """Subscribe notification channels to a wallet."""
    data = request.get_json()
    wallet = data.get('wallet', '').strip()
    channels = data.get('channels', [])
    
    if not wallet:
        return jsonify({"error": "Wallet required"}), 400
    
    if not channels:
        return jsonify({"error": "At least one channel required"}), 400
    
    result = subscribe_to_wallet(wallet, channels)
    return jsonify(result)


@app.route('/api/notifications/unsubscribe', methods=['POST'])
def unsubscribe_wallet():
    """Unsubscribe notification channels from a wallet."""
    data = request.get_json()
    wallet = data.get('wallet', '').strip()
    channels = data.get('channels')
    
    if not wallet:
        return jsonify({"error": "Wallet required"}), 400
    
    result = unsubscribe_from_wallet(wallet, channels)
    if result.get("error"):
        return jsonify(result), 404
    return jsonify(result)


@app.route('/api/notifications/test', methods=['POST'])
def test_notification():
    """Send a test notification to a channel."""
    data = request.get_json()
    channel_type = data.get('type', '').strip()
    channel_name = data.get('name', '').strip()
    
    if not channel_type or not channel_name:
        return jsonify({"error": "Type and name required"}), 400
    
    config = get_notification_channels()
    
    if channel_type == 'discord':
        from notifications import send_discord_notification, format_alert_discord_embed, load_notifications_config
        full_config = load_notifications_config()
        webhook = full_config.get('discord_webhooks', {}).get(channel_name, {})
        if not webhook:
            return jsonify({"error": "Discord webhook not found"}), 404
        
        embed = format_alert_discord_embed(
            "Test Notification",
            "This is a test notification from PolyClaw! 🦞\nYour Discord integration is working correctly.",
            "info"
        )
        success = send_discord_notification(webhook['url'], embed)
        return jsonify({"success": success})
    
    elif channel_type == 'telegram':
        from notifications import send_telegram_notification, format_alert_telegram_message, load_notifications_config
        full_config = load_notifications_config()
        tg = full_config.get('telegram_configs', {}).get(channel_name, {})
        if not tg:
            return jsonify({"error": "Telegram config not found"}), 404
        
        message = format_alert_telegram_message(
            "Test Notification",
            "This is a test notification from PolyClaw! 🦞\nYour Telegram integration is working correctly.",
            "info"
        )
        success = send_telegram_notification(tg['bot_token'], tg['chat_id'], message)
        return jsonify({"success": success})
    
    return jsonify({"error": "Invalid channel type"}), 400


@app.route('/api/terminal/<wallet>', methods=['GET', 'OPTIONS'])
def terminal_data(wallet):
    """Fetch trades + full terminal analysis for the wallet analyzer."""
    # Handle CORS preflight
    if request.method == 'OPTIONS':
        resp = app.make_default_options_response()
        resp.headers['Access-Control-Allow-Origin'] = '*'
        resp.headers['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return resp

    limit = request.args.get('limit', 200, type=int)
    result = fetch_trades_with_limit(wallet, min(limit, 500))
    trades = result.get('trades', [])

    # Run full analysis pipeline
    base_analysis = {}
    terminal = {}
    if trades:
        base_analysis = analyze_trades(trades)
        terminal = run_terminal_analysis(trades, base_analysis)

    # Slim trade objects for the feed
    slim_trades = []
    for t in trades[:100]:  # Cap feed at 100
        slim_trades.append({
            'ts': t.get('timestamp'),
            's': t.get('side'),
            'p': t.get('price'),
            'sz': t.get('size'),
            'v': t.get('usdcSize'),
            't': (t.get('title') or '')[:80],
            'o': t.get('outcome'),
            'tx': (t.get('transactionHash') or '')[:12],
        })

    # Extract key data from base analysis
    pnl = base_analysis.get('pnl', {}).get('summary', {})
    summary = base_analysis.get('summary', {})
    freq = base_analysis.get('frequency', {})

    payload = {
        "w": wallet,
        "u": result.get('username'),
        "n": len(trades),
        "trades": slim_trades,
        # Base metrics
        "pnl": pnl,
        "summary": summary,
        "freq": {
            "sub_second_pct": freq.get('gaps', {}).get('sub_second_percentage', 0),
            "is_bot": freq.get('is_bot_like', False),
            "bursts": freq.get('bursts', {}).get('total_bursts', 0),
        },
        # Terminal-specific metrics
        "kelly": terminal.get('kelly', {}),
        "risk": terminal.get('risk_adjusted', {}),
        "hc": terminal.get('high_confidence', {}),
        "flow": terminal.get('order_flow', {}),
        "phases": terminal.get('phases', {}),
        "exec": terminal.get('execution', {}),
        "tpnl": terminal.get('time_pnl', {}),
        "pos": terminal.get('positions', []),
        "fp": terminal.get('fingerprint', {}),
    }

    # JSONP support for proxy environments
    callback = request.args.get('callback')
    if callback:
        js = f"{callback}({json.dumps(payload)})"
        resp = Response(js, mimetype='application/javascript')
    else:
        resp = jsonify(payload)

    resp.headers['Access-Control-Allow-Origin'] = '*'
    return resp


@app.route('/api/trades/<wallet>')
def get_trades(wallet):
    """Fetch trades - use mode=full for all history."""
    mode = request.args.get('mode', 'recent')
    limit = request.args.get('limit', 100, type=int)
    start_date = request.args.get('startDate')
    end_date = request.args.get('endDate')

    if mode == 'full':
        result = fetch_all_trades(wallet)
    elif limit > 1000:
        # Use paginated fetch for larger limits (5k, 10k, 50k)
        result = fetch_trades_with_limit(wallet, limit)
    else:
        result = fetch_recent_trades(wallet, limit)

    # Apply date filtering if provided
    if start_date or end_date:
        result = filter_trades_by_date(result, start_date, end_date)

    # Save to file
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    short_wallet = wallet[:10]
    suffix = '-full' if mode == 'full' else ''

    with open(f"{OUTPUT_DIR}/{short_wallet}{suffix}.json", 'w') as f:
        json.dump(result, f, indent=2)

    return jsonify(result)


@app.route('/api/trades/<wallet>/stream')
def stream_trades(wallet):
    """
    Stream trade fetching progress via Server-Sent Events (SSE).
    Use this for large fetches to show real-time progress.
    """
    limit = request.args.get('limit', type=int)
    
    def generate():
        yield from fetch_trades_stream(wallet, limit)
    
    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'X-Accel-Buffering': 'no'
        }
    )


@app.route('/api/download/<wallet>/<format>')
def download_trades(wallet, format):
    """Download trades as JSON, CSV, or TXT."""
    mode = request.args.get('mode', 'recent')
    limit = request.args.get('limit', 100, type=int)

    if mode == 'full':
        result = fetch_all_trades(wallet)
    elif limit > 1000:
        # Use paginated fetch for larger limits
        result = fetch_trades_with_limit(wallet, limit)
    else:
        result = fetch_recent_trades(wallet, limit)

    short_wallet = wallet[:10]
    username = result.get('username') or short_wallet
    suffix = '-full' if mode == 'full' else ''

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if format == 'json':
        filename = f"{OUTPUT_DIR}/{username}{suffix}-trades.json"
        with open(filename, 'w') as f:
            json.dump(result, f, indent=2)
        return send_file(filename, as_attachment=True, download_name=f"{username}{suffix}-trades.json")

    elif format == 'csv':
        output = io.StringIO()
        if result['trades']:
            fieldnames = ['timestamp', 'datetime', 'type', 'side', 'price', 'size',
                         'usdcSize', 'title', 'outcome', 'transactionHash']
            writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()

            for trade in result['trades']:
                ts = trade.get('timestamp', 0)
                trade['datetime'] = datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S') if ts else ''
                writer.writerow(trade)

        csv_filename = f"{OUTPUT_DIR}/{username}{suffix}-trades.csv"
        with open(csv_filename, 'w') as f:
            f.write(output.getvalue())

        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode()),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f"{username}{suffix}-trades.csv"
        )

    elif format == 'txt':
        # Generate formatted text file
        lines = []
        lines.append("=" * 60)
        lines.append("POLYMARKET TRADE HISTORY")
        lines.append("=" * 60)
        lines.append("")
        lines.append(f"Wallet: {wallet}")
        lines.append(f"Username: {username}")
        lines.append(f"Total Trades: {len(result.get('trades', []))}")
        lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("")
        lines.append("-" * 60)
        lines.append("")

        for i, trade in enumerate(result.get('trades', []), 1):
            ts = trade.get('timestamp', 0)
            dt = datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S') if ts else 'N/A'
            side = trade.get('side', 'N/A')
            price = trade.get('price', 0)
            size = trade.get('size', 0)
            usdc = trade.get('usdcSize', 0)
            title = trade.get('title', 'N/A')
            outcome = trade.get('outcome', 'N/A')
            tx_hash = trade.get('transactionHash', 'N/A')

            lines.append(f"Trade #{i}")
            lines.append(f"  Time:     {dt}")
            lines.append(f"  Side:     {side}")
            lines.append(f"  Market:   {title}")
            lines.append(f"  Outcome:  {outcome}")
            lines.append(f"  Price:    {price:.4f}" if isinstance(price, (int, float)) else f"  Price:    {price}")
            lines.append(f"  Size:     {size:.2f}" if isinstance(size, (int, float)) else f"  Size:     {size}")
            lines.append(f"  Value:    ${usdc:.2f}" if isinstance(usdc, (int, float)) else f"  Value:    ${usdc}")
            lines.append(f"  Tx:       {tx_hash[:20]}..." if tx_hash and len(str(tx_hash)) > 20 else f"  Tx:       {tx_hash}")
            lines.append("")

        txt_content = "\n".join(lines)
        txt_filename = f"{OUTPUT_DIR}/{username}{suffix}-trades.txt"
        with open(txt_filename, 'w') as f:
            f.write(txt_content)

        return send_file(
            io.BytesIO(txt_content.encode()),
            mimetype='text/plain',
            as_attachment=True,
            download_name=f"{username}{suffix}-trades.txt"
        )

    elif format == 'xlsx':
        # Generate Excel file using openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return jsonify({"error": "Excel export not available. Install openpyxl."}), 500

        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = ['#', 'Date/Time', 'Side', 'Market', 'Outcome', 'Price', 'Size', 'Value (USDC)', 'Transaction Hash']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        for row_num, trade in enumerate(result.get('trades', []), 2):
            ts = trade.get('timestamp', 0)
            dt = datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S') if ts else 'N/A'
            
            row_data = [
                row_num - 1,
                dt,
                trade.get('side', 'N/A'),
                trade.get('title', 'N/A'),
                trade.get('outcome', 'N/A'),
                trade.get('price', 0),
                trade.get('size', 0),
                trade.get('usdcSize', 0),
                trade.get('transactionHash', 'N/A')
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                if col in [6, 7, 8]:  # Price, Size, Value columns
                    cell.number_format = '#,##0.0000' if col == 6 else '#,##0.00'

        # Adjust column widths
        column_widths = [6, 20, 8, 50, 15, 12, 12, 15, 45]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Save to BytesIO
        excel_output = io.BytesIO()
        wb.save(excel_output)
        excel_output.seek(0)

        return send_file(
            excel_output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{username}{suffix}-trades.xlsx"
        )

    return jsonify({"error": "Invalid format. Use 'json', 'csv', 'txt', or 'xlsx'"}), 400


@app.route('/api/check/<wallet>')
def check_existing(wallet):
    """Check if we have existing data for a wallet."""
    short_wallet = wallet[:10]

    files = []
    for suffix in ['', '-full']:
        for ext in ['json', 'csv']:
            path = f"{OUTPUT_DIR}/{short_wallet}{suffix}.{ext}"
            if os.path.exists(path):
                size = os.path.getsize(path)
                files.append({"path": path, "size": size})

    return jsonify({"wallet": wallet, "files": files})


@app.route('/api/analyze/<wallet>')
def analyze_wallet(wallet):
    """Analyze trades and return pattern insights."""
    limit = request.args.get('limit', 1000, type=int)
    mode = request.args.get('mode', 'recent')
    start_date = request.args.get('startDate')
    end_date = request.args.get('endDate')

    # Fetch trades
    if mode == 'full':
        result = fetch_all_trades(wallet)
    else:
        # Use limit for scalable fetching
        result = fetch_trades_with_limit(wallet, limit)

    # Apply date filtering if provided
    if start_date or end_date:
        result = filter_trades_by_date(result, start_date, end_date)

    trades = result.get('trades', [])
    if not trades:
        return jsonify({"error": "No trades found", "wallet": wallet})

    # Run analysis
    analysis = analyze_trades(trades)

    return jsonify({
        "wallet": wallet,
        "username": result.get('username'),
        "trade_count": len(trades),
        "analysis": analysis
    })


@app.route('/api/analyze/<wallet>/report')
def get_analysis_report(wallet):
    """Generate and download markdown analysis report."""
    limit = request.args.get('limit', 1000, type=int)
    mode = request.args.get('mode', 'recent')

    # Fetch trades
    if mode == 'full':
        result = fetch_all_trades(wallet)
    else:
        result = fetch_trades_with_limit(wallet, limit)

    trades = result.get('trades', [])
    if not trades:
        return jsonify({"error": "No trades found"}), 404

    # Run analysis
    analysis = analyze_trades(trades)

    # Generate report
    report = generate_report(analysis, wallet, result.get('username'))

    # Return as downloadable markdown
    username = result.get('username') or wallet[:10]
    return Response(
        report,
        mimetype='text/markdown',
        headers={'Content-Disposition': f'attachment; filename={username}-analysis.md'}
    )


# =============================================================================
# AI ANALYSIS ENDPOINTS
# =============================================================================

@app.route('/api/ai-analyze/<wallet>', methods=['POST'])
def ai_analyze_wallet(wallet):
    """
    AI-powered trade analysis using OpenAI or Anthropic.
    
    Request body:
    {
        "provider": "openai" | "anthropic",
        "prompt_type": "strategy" | "risk" | "performance" | "custom",
        "custom_prompt": "optional custom question",
        "limit": 1000
    }
    """
    data = request.get_json() or {}
    provider = data.get('provider', 'openai')
    prompt_type = data.get('prompt_type', 'strategy')
    custom_prompt = data.get('custom_prompt')
    limit = data.get('limit', 1000)
    
    # Validate provider
    if provider not in ['openai', 'anthropic']:
        return jsonify({"error": "Invalid provider. Use 'openai' or 'anthropic'"}), 400
    
    # Validate prompt type
    valid_types = ['strategy', 'risk', 'performance', 'custom']
    if prompt_type not in valid_types:
        return jsonify({"error": f"Invalid prompt_type. Use one of: {valid_types}"}), 400
    
    # Fetch trades
    result = fetch_trades_with_limit(wallet, limit)
    trades = result.get('trades', [])
    
    if not trades:
        return jsonify({"error": "No trades found for this wallet"}), 404
    
    # Run pattern analysis first (for context)
    analysis = analyze_trades(trades)
    
    # Run AI analysis
    ai_result = run_ai_analysis(
        trades=trades,
        analysis=analysis,
        provider=provider,
        prompt_type=prompt_type,
        custom_prompt=custom_prompt
    )
    
    if ai_result.get('error'):
        return jsonify(ai_result), 500
    
    # Add wallet info
    ai_result['wallet'] = wallet
    ai_result['username'] = result.get('username')
    
    return jsonify(ai_result)


@app.route('/api/ai-providers')
def get_ai_providers():
    """Check which AI providers are available and configured."""
    providers = get_available_providers()
    return jsonify(providers)


@app.route('/api/chat', methods=['POST'])
def chat_with_polyclaw():
    """PolyClaw chat endpoint - conversational AI for Polymarket trading."""
    data = request.get_json()
    message = data.get('message', '')
    history = data.get('history', [])
    context = data.get('context', {})
    
    if not message:
        return jsonify({"error": "No message provided"}), 400
    
    # Check for available providers
    providers = get_available_providers()
    provider = None
    if providers.get('anthropic', {}).get('available'):
        provider = 'anthropic'
    elif providers.get('openai', {}).get('available'):
        provider = 'openai'
    
    if not provider:
        return jsonify({
            "error": "No AI provider configured",
            "response": "I need an AI provider to chat! Please set OPENAI_API_KEY or ANTHROPIC_API_KEY in your .env file."
        })
    
    # Build system prompt for PolyClaw
    system_prompt = """You are PolyClaw 🦞, an AI trading assistant specialized in Polymarket prediction markets.

Your expertise includes:
- Analyzing wallet trading strategies and patterns
- Helping users understand copy trading and find profitable traders
- Building trading bots and automation strategies
- Finding arbitrage opportunities across markets
- Risk management and position sizing (Kelly Criterion)
- Market analysis and sentiment interpretation

Personality:
- Friendly and approachable, but data-driven
- You use 🦞 occasionally but not excessively  
- You give specific, actionable advice
- You explain complex concepts simply
- You're honest about risks and uncertainties

Current context:
"""
    
    if context.get('current_wallet'):
        system_prompt += f"\n- User is analyzing wallet: {context['current_wallet']}"
    if context.get('trades_loaded'):
        system_prompt += f"\n- {context['trades_loaded']} trades loaded"
    if context.get('summary'):
        summary = context['summary']
        system_prompt += f"\n- Summary stats: {summary.get('total_trades', 'N/A')} trades, {summary.get('win_rate', 'N/A')}% win rate, ${summary.get('net_pnl', 'N/A')} net P&L"
    
    system_prompt += "\n\nBe helpful, specific, and data-driven in your responses. If you don't have enough information, ask clarifying questions."
    
    # Build messages for API
    messages = [{"role": "system", "content": system_prompt}]
    
    # Add recent history (limited to avoid token limits)
    for msg in history[-8:]:
        if msg.get('role') in ['user', 'assistant']:
            messages.append({
                "role": msg['role'],
                "content": msg['content']
            })
    
    # Add current message
    messages.append({"role": "user", "content": message})
    
    try:
        if provider == 'anthropic':
            import anthropic
            client = anthropic.Anthropic(api_key=os.environ.get('ANTHROPIC_API_KEY'))
            response = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=1024,
                system=system_prompt,
                messages=[m for m in messages if m['role'] != 'system']
            )
            reply = response.content[0].text
        else:
            import openai
            client = openai.OpenAI(api_key=os.environ.get('OPENAI_API_KEY'))
            response = client.chat.completions.create(
                model="gpt-4o",
                messages=messages,
                max_tokens=1024
            )
            reply = response.choices[0].message.content
        
        return jsonify({"response": reply, "provider": provider})
    
    except Exception as e:
        return jsonify({
            "error": str(e),
            "response": f"Sorry, I had trouble processing that. Error: {str(e)}"
        }), 500


def fetch_trades_with_limit(wallet: str, limit: int) -> Dict:
    """Fetch trades up to a specific limit using pagination."""
    all_trades = []
    seen_keys: Set[tuple] = set()
    end_ts = None
    username = None

    while len(all_trades) < limit:
        batch_limit = min(BATCH_SIZE, limit - len(all_trades))
        params = {"user": wallet, "limit": batch_limit}
        if end_ts:
            params["end"] = end_ts

        try:
            response = requests.get(f"{DATA_API}/activity", params=params, timeout=30, proxies={"http": None, "https": None})
            response.raise_for_status()
            trades = response.json()
        except Exception as e:
            break

        if not trades or not isinstance(trades, list):
            break

        if not username and trades[0].get('name'):
            username = trades[0].get('name')

        oldest_ts = float('inf')
        new_count = 0

        for t in trades:
            if len(all_trades) >= limit:
                break
            key = (t.get('transactionHash'), t.get('timestamp'), t.get('asset'))
            if key not in seen_keys:
                seen_keys.add(key)
                all_trades.append(t)
                new_count += 1

            ts = t.get('timestamp', 0)
            if ts and ts < oldest_ts:
                oldest_ts = ts

        if len(trades) < batch_limit or new_count == 0:
            break

        if oldest_ts != float('inf'):
            end_ts = oldest_ts
        else:
            break

        time.sleep(0.1)

    all_trades.sort(key=lambda t: t.get('timestamp', 0), reverse=True)

    return {
        "wallet": wallet,
        "username": username,
        "fetched_at": datetime.now(timezone.utc).isoformat(),
        "trade_count": len(all_trades),
        "trades": all_trades
    }


# Trade count presets for UI
TRADE_PRESETS = [100, 1000, 5000, 10000, 50000, 100000]


@app.route('/api/presets')
def get_presets():
    """Return available trade count presets."""
    return jsonify({"presets": TRADE_PRESETS})


# =============================================================================
# COMPARISON MODE
# =============================================================================

@app.route('/api/compare', methods=['POST'])
def compare_wallets_endpoint():
    """Compare multiple wallets side-by-side."""
    data = request.get_json()
    wallets = data.get('wallets', [])
    limit = data.get('limit', 1000)
    tier = data.get('tier', 'free')

    if not wallets:
        return jsonify({"error": "No wallets provided"}), 400

    # Check tier access
    tier_info = get_tier_info(tier)
    max_wallets = tier_info.get("comparison_wallets", 0)

    if len(wallets) > max_wallets and tier != 'scale':
        return jsonify({
            "error": f"Your plan allows comparing up to {max_wallets} wallets. Upgrade to compare more.",
            "tier": tier,
            "max_wallets": max_wallets,
            "upgrade_needed": True
        }), 403

    # Fetch and analyze each wallet
    wallet_analyses = []
    for wallet in wallets[:max_wallets or len(wallets)]:
        result = fetch_trades_with_limit(wallet, limit)
        trades = result.get('trades', [])

        if trades:
            analysis = analyze_trades(trades)
            wallet_analyses.append({
                "wallet": wallet,
                "username": result.get('username'),
                "trade_count": len(trades),
                "analysis": analysis,
                "pnl": analysis.get("pnl", {})
            })

    # Run comparison
    comparison = compare_wallets(wallet_analyses)

    return jsonify({
        "comparison": comparison,
        "wallet_data": wallet_analyses,
        "tier": tier
    })


# =============================================================================
# PRICING ENDPOINTS
# =============================================================================

@app.route('/api/pricing')
def get_pricing():
    """Return pricing tier information."""
    return jsonify({
        "tiers": PRICING_TIERS,
        "overage_rate": 0.01,
        "currency": "USD"
    })


@app.route('/api/pricing/<tier>')
def get_tier(tier):
    """Get specific tier information."""
    tier_info = get_tier_info(tier)
    if tier_info:
        return jsonify(tier_info)
    return jsonify({"error": "Invalid tier"}), 404


@app.route('/api/usage/check', methods=['POST'])
def check_usage():
    """Check if usage is within tier limits."""
    data = request.get_json()
    tier = data.get('tier', 'free')
    trades_used = data.get('trades_used', 0)

    tier_info = get_tier_info(tier)
    limit = tier_info.get("trades_per_month", 100)
    overage = calculate_overage(tier, trades_used)

    return jsonify({
        "tier": tier,
        "limit": limit,
        "used": trades_used,
        "remaining": max(0, limit - trades_used),
        "percentage_used": round(trades_used / limit * 100, 1),
        "overage": overage,
        "within_limit": trades_used <= limit
    })


@app.route('/api/features/<feature>')
def check_feature(feature):
    """Check which tiers have access to a feature."""
    tiers_with_access = []
    for tier_id, tier_info in PRICING_TIERS.items():
        if feature in tier_info.get("features", []):
            tiers_with_access.append({
                "tier": tier_id,
                "name": tier_info["name"],
                "price": tier_info["price"]
            })

    return jsonify({
        "feature": feature,
        "available_in": tiers_with_access,
        "minimum_tier": tiers_with_access[0]["tier"] if tiers_with_access else None
    })


# =============================================================================
# POLYCLAW STRATEGY ENGINE ENDPOINTS
# =============================================================================

@app.route('/api/strategy/templates')
def get_strategy_templates():
    """Get all available strategy templates for ideation."""
    return jsonify({
        "templates": STRATEGY_TEMPLATES,
        "ideation_prompts": IDEATION_PROMPTS
    })


@app.route('/api/strategy/diagnose/<wallet>')
def diagnose_wallet_strategy(wallet):
    """Analyze and diagnose the trading strategy of a wallet."""
    limit = request.args.get('limit', 1000, type=int)
    
    # Fetch trades
    result = fetch_trades_with_limit(wallet, limit)
    trades = result.get('trades', [])
    
    if not trades:
        return jsonify({"error": "No trades found", "wallet": wallet}), 404
    
    # Run strategy diagnosis
    diagnosis = analyze_wallet_strategy(trades, wallet)
    diagnosis["username"] = result.get("username")
    diagnosis["trade_count"] = len(trades)
    
    return jsonify(diagnosis)


@app.route('/api/strategy/ideate', methods=['POST'])
def ideate_strategies():
    """Generate strategy ideas based on parameters."""
    data = request.get_json() or {}
    
    wallet = data.get('wallet')
    wallet_analysis = None
    
    # If wallet provided, include its analysis
    if wallet:
        result = fetch_trades_with_limit(wallet, 1000)
        trades = result.get('trades', [])
        if trades:
            wallet_analysis = analyze_wallet_strategy(trades, wallet)
    
    ideas = generate_strategy_ideas(
        wallet_analysis=wallet_analysis,
        market_category=data.get('market_category'),
        risk_tolerance=data.get('risk_tolerance', 'medium'),
        capital=data.get('capital', 10000)
    )
    
    return jsonify(ideas)


@app.route('/api/strategy/define', methods=['POST'])
def create_strategy():
    """Define a new trading strategy."""
    data = request.get_json()
    
    if not data.get('name'):
        return jsonify({"error": "Strategy name required"}), 400
    if not data.get('entry_rules'):
        return jsonify({"error": "Entry rules required"}), 400
    
    strategy = define_strategy(
        name=data['name'],
        entry_rules=data.get('entry_rules', []),
        exit_rules=data.get('exit_rules', []),
        position_sizing=data.get('position_sizing', '5% of bankroll'),
        risk_params=data.get('risk_params')
    )
    
    return jsonify(strategy)


@app.route('/api/strategy/iterate', methods=['POST'])
def iterate_existing_strategy():
    """Iterate and improve a strategy based on performance."""
    data = request.get_json()
    
    if not data.get('strategy'):
        return jsonify({"error": "Current strategy required"}), 400
    
    iteration = iterate_strategy(
        current_strategy=data['strategy'],
        performance_data=data.get('performance', {}),
        feedback=data.get('feedback')
    )
    
    return jsonify(iteration)


# =============================================================================
# REAL-TIME WALLET TRACKING
# =============================================================================

@app.route('/api/track/<wallet>/start', methods=['POST'])
def start_tracking(wallet):
    """Start real-time tracking of a wallet."""
    if wallet in tracked_wallets:
        return jsonify({"status": "already_tracking", "wallet": wallet})
    
    # Initial fetch
    result = fetch_recent_trades(wallet, 100)
    
    tracked_wallets[wallet] = {
        "wallet": wallet,
        "username": result.get("username"),
        "started_at": datetime.now(timezone.utc).isoformat(),
        "last_check": datetime.now(timezone.utc).isoformat(),
        "trade_count": result.get("trade_count", 0),
        "latest_trade_ts": result['trades'][0].get('timestamp') if result.get('trades') else None,
        "new_trades": [],
        "status": "active"
    }
    
    return jsonify({
        "status": "tracking_started",
        "wallet": wallet,
        "username": result.get("username"),
        "current_trades": result.get("trade_count", 0)
    })


@app.route('/api/track/<wallet>/stop', methods=['POST'])
def stop_tracking(wallet):
    """Stop tracking a wallet."""
    if wallet in tracked_wallets:
        del tracked_wallets[wallet]
        return jsonify({"status": "tracking_stopped", "wallet": wallet})
    return jsonify({"status": "not_tracking", "wallet": wallet})


@app.route('/api/track/<wallet>/status')
def get_tracking_status(wallet):
    """Get current tracking status and any new trades."""
    if wallet not in tracked_wallets:
        return jsonify({"status": "not_tracking", "wallet": wallet})
    
    tracking_data = tracked_wallets[wallet]
    
    # Check for new trades
    result = fetch_recent_trades(wallet, 50)
    new_trades = []
    
    if result.get('trades') and tracking_data.get('latest_trade_ts'):
        for trade in result['trades']:
            if trade.get('timestamp', 0) > tracking_data['latest_trade_ts']:
                new_trades.append(trade)
    
    # Update tracking data
    if new_trades:
        tracking_data['new_trades'] = new_trades
        tracking_data['latest_trade_ts'] = result['trades'][0].get('timestamp')
        tracking_data['trade_count'] += len(new_trades)
    
    tracking_data['last_check'] = datetime.now(timezone.utc).isoformat()
    
    return jsonify({
        "status": "tracking",
        "wallet": wallet,
        "username": tracking_data.get("username"),
        "started_at": tracking_data.get("started_at"),
        "last_check": tracking_data.get("last_check"),
        "total_trades": tracking_data.get("trade_count"),
        "new_trades_count": len(new_trades),
        "new_trades": new_trades
    })


@app.route('/api/track/<wallet>/stream')
def stream_wallet_updates(wallet):
    """Stream real-time updates for a tracked wallet via SSE."""
    def generate():
        last_ts = None
        
        # Initial status
        if wallet in tracked_wallets:
            last_ts = tracked_wallets[wallet].get('latest_trade_ts')
        else:
            # Auto-start tracking
            result = fetch_recent_trades(wallet, 50)
            if result.get('trades'):
                last_ts = result['trades'][0].get('timestamp')
            
            yield f"data: {json.dumps({'type': 'connected', 'wallet': wallet})}\n\n"
        
        check_count = 0
        while True:
            check_count += 1
            
            try:
                result = fetch_recent_trades(wallet, 20)
                new_trades = []
                
                if result.get('trades'):
                    for trade in result['trades']:
                        trade_ts = trade.get('timestamp', 0)
                        if last_ts and trade_ts > last_ts:
                            new_trades.append(trade)
                    
                    if new_trades:
                        last_ts = result['trades'][0].get('timestamp')
                        yield f"data: {json.dumps({'type': 'new_trades', 'count': len(new_trades), 'trades': new_trades})}\n\n"
                
                # Heartbeat every 10 checks
                if check_count % 10 == 0:
                    yield f"data: {json.dumps({'type': 'heartbeat', 'check': check_count})}\n\n"
                
            except Exception as e:
                yield f"data: {json.dumps({'type': 'error', 'message': str(e)})}\n\n"
            
            time.sleep(10)  # Check every 10 seconds
    
    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'X-Accel-Buffering': 'no'
        }
    )


@app.route('/api/track/list')
def list_tracked_wallets():
    """List all currently tracked wallets."""
    return jsonify({
        "tracked_wallets": list(tracked_wallets.keys()),
        "count": len(tracked_wallets),
        "details": [
            {
                "wallet": w,
                "username": d.get("username"),
                "started_at": d.get("started_at"),
                "trade_count": d.get("trade_count")
            }
            for w, d in tracked_wallets.items()
        ]
    })


# ============================================================
# DAEMON CONTROL ENDPOINTS
# ============================================================

DAEMON_PID_FILE = os.path.expanduser("~/.polyclaw/daemon.pid")
TRACKING_FILE = os.path.expanduser("~/.polyclaw/tracking.json")


def load_cli_tracking():
    """Load tracked wallets from CLI tracking file."""
    if os.path.exists(TRACKING_FILE):
        try:
            with open(TRACKING_FILE) as f:
                return json.load(f)
        except:
            pass
    return {"wallets": []}


def save_cli_tracking(data):
    """Save tracked wallets to CLI tracking file."""
    os.makedirs(os.path.dirname(TRACKING_FILE), exist_ok=True)
    with open(TRACKING_FILE, 'w') as f:
        json.dump(data, f, indent=2)


@app.route('/api/daemon/status')
def daemon_status():
    """Get daemon status."""
    import signal
    
    if not os.path.exists(DAEMON_PID_FILE):
        return jsonify({
            "running": False,
            "message": "Daemon is not running"
        })
    
    try:
        with open(DAEMON_PID_FILE) as f:
            pid = int(f.read().strip())
        
        # Check if process is running
        os.kill(pid, 0)
        
        # Load tracking info
        tracking = load_cli_tracking()
        
        return jsonify({
            "running": True,
            "pid": pid,
            "tracked_wallets": tracking.get("wallets", []),
            "tracked_count": len(tracking.get("wallets", []))
        })
    except (OSError, ValueError):
        # Process not running, clean up stale PID file
        try:
            os.unlink(DAEMON_PID_FILE)
        except:
            pass
        return jsonify({
            "running": False,
            "message": "Daemon is not running (cleaned stale PID)"
        })


@app.route('/api/daemon/start', methods=['POST'])
def daemon_start():
    """Start the background daemon."""
    import subprocess
    import signal
    
    # Check if already running
    if os.path.exists(DAEMON_PID_FILE):
        try:
            with open(DAEMON_PID_FILE) as f:
                pid = int(f.read().strip())
            os.kill(pid, 0)
            return jsonify({
                "success": False,
                "error": f"Daemon is already running (PID: {pid})"
            })
        except (OSError, ValueError):
            try:
                os.unlink(DAEMON_PID_FILE)
            except:
                pass
    
    # Start daemon in background
    try:
        daemon_path = os.path.join(os.path.dirname(__file__), 'daemon.py')
        
        # Start as background process
        process = subprocess.Popen(
            ['python', daemon_path],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            start_new_session=True
        )
        
        # Give it a moment to start and write PID
        time.sleep(1)
        
        if os.path.exists(DAEMON_PID_FILE):
            with open(DAEMON_PID_FILE) as f:
                pid = int(f.read().strip())
            return jsonify({
                "success": True,
                "message": f"Daemon started (PID: {pid})",
                "pid": pid
            })
        else:
            return jsonify({
                "success": True,
                "message": "Daemon starting...",
                "pid": process.pid
            })
            
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        })


@app.route('/api/daemon/stop', methods=['POST'])
def daemon_stop():
    """Stop the background daemon."""
    import signal
    
    if not os.path.exists(DAEMON_PID_FILE):
        return jsonify({
            "success": False,
            "error": "Daemon is not running"
        })
    
    try:
        with open(DAEMON_PID_FILE) as f:
            pid = int(f.read().strip())
        
        os.kill(pid, signal.SIGTERM)
        os.unlink(DAEMON_PID_FILE)
        
        return jsonify({
            "success": True,
            "message": f"Daemon stopped (was PID: {pid})"
        })
    except OSError:
        try:
            os.unlink(DAEMON_PID_FILE)
        except:
            pass
        return jsonify({
            "success": True,
            "message": "Daemon was not running (cleaned up)"
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        })


@app.route('/api/tracking/list')
def get_tracking_list():
    """Get list of tracked wallets from CLI tracking file."""
    tracking = load_cli_tracking()
    wallets = tracking.get("wallets", [])
    
    # Get basic info for each wallet
    wallet_details = []
    for wallet in wallets:
        wallet_details.append({
            "address": wallet,
            "short": f"{wallet[:6]}...{wallet[-4:]}" if len(wallet) > 12 else wallet
        })
    
    return jsonify({
        "wallets": wallet_details,
        "count": len(wallets)
    })


@app.route('/api/tracking/add', methods=['POST'])
def add_tracking():
    """Add a wallet to tracking list."""
    data = request.json
    wallet = data.get("wallet", "").strip()
    
    if not wallet:
        return jsonify({"success": False, "error": "Wallet address required"})
    
    tracking = load_cli_tracking()
    
    if wallet in tracking["wallets"]:
        return jsonify({"success": False, "error": "Wallet already tracked"})
    
    tracking["wallets"].append(wallet)
    save_cli_tracking(tracking)
    
    return jsonify({
        "success": True,
        "message": f"Now tracking {wallet[:8]}...",
        "count": len(tracking["wallets"])
    })


@app.route('/api/tracking/remove', methods=['POST'])
def remove_tracking():
    """Remove a wallet from tracking list."""
    data = request.json
    wallet = data.get("wallet", "").strip()
    
    if not wallet:
        return jsonify({"success": False, "error": "Wallet address required"})
    
    tracking = load_cli_tracking()
    
    if wallet not in tracking["wallets"]:
        return jsonify({"success": False, "error": "Wallet not tracked"})
    
    tracking["wallets"].remove(wallet)
    save_cli_tracking(tracking)
    
    return jsonify({
        "success": True,
        "message": f"Stopped tracking {wallet[:8]}...",
        "count": len(tracking["wallets"])
    })


@app.route('/api/system/info')
def system_info():
    """Get system information for control panel."""
    import platform
    
    # Get daemon status
    daemon_running = False
    daemon_pid = None
    if os.path.exists(DAEMON_PID_FILE):
        try:
            with open(DAEMON_PID_FILE) as f:
                daemon_pid = int(f.read().strip())
            os.kill(daemon_pid, 0)
            daemon_running = True
        except:
            pass
    
    # Get tracking info
    tracking = load_cli_tracking()
    
    # Get notification channels
    try:
        channels = get_notification_channels()
        discord_count = len(channels.get("discord", {}))
        telegram_count = len(channels.get("telegram", {}))
    except:
        discord_count = 0
        telegram_count = 0
    
    # Get AI providers
    try:
        providers = get_available_providers()
    except:
        providers = []
    
    return jsonify({
        "daemon": {
            "running": daemon_running,
            "pid": daemon_pid
        },
        "tracking": {
            "count": len(tracking.get("wallets", [])),
            "wallets": tracking.get("wallets", [])[:5]  # First 5
        },
        "channels": {
            "discord": discord_count,
            "telegram": telegram_count
        },
        "ai": {
            "providers": providers,
            "available": len(providers) > 0
        },
        "system": {
            "platform": platform.system(),
            "python": platform.python_version()
        }
    })


if __name__ == '__main__':
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    RED = '\033[91m'
    DARK_RED = '\033[31m'
    GRAY = '\033[90m'
    RESET = '\033[0m'
    BOLD = '\033[1m'
    
    print(f"\n{RED}{BOLD}🦞 PolyClaw{RESET} {GRAY}- Trading Strategy Platform{RESET}")
    print(f"   {DARK_RED}http://localhost:8080{RESET}")
    print(f"\n   {RED}📊 Trade Data:{RESET}")
    print(f"   {GRAY}   /api/trades/{{wallet}}{RESET}")
    print(f"   {GRAY}   /api/trades/{{wallet}}?mode=full{RESET}")
    print(f"\n   {RED}🧠 Strategy Engine:{RESET}")
    print(f"   {GRAY}   /api/strategy/templates{RESET}")
    print(f"   {GRAY}   /api/strategy/diagnose/{{wallet}}{RESET}")
    print(f"   {GRAY}   /api/strategy/ideate{RESET}")
    print(f"\n   {RED}📡 Real-time Tracking:{RESET}")
    print(f"   {GRAY}   /api/track/{{wallet}}/start{RESET}")
    print(f"   {GRAY}   /api/track/{{wallet}}/stream{RESET}")
    print()
    app.run(debug=True, port=8080, host='0.0.0.0')
