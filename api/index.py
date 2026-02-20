#!/usr/bin/env python3
"""
PolyEdge.io - Vercel Serverless Entry Point
"""

import sys
import os

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from flask import Flask, jsonify, request, send_from_directory, send_file, Response
import requests
import json
import csv
import io
import time
from datetime import datetime, timezone
from typing import List, Dict, Set, Optional

# Initialize Flask app
app = Flask(__name__, static_folder='../')

# Configuration
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:////tmp/polyedge.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Initialize database and auth (optional - may fail on serverless without proper DB)
DB_AVAILABLE = False
try:
    from models import db, init_db, User
    init_db(app)
    from auth import init_auth, login_required, current_user
    init_auth(app)
    from payments import init_payments
    init_payments(app)
    DB_AVAILABLE = True
except Exception as e:
    print(f"DB/Auth initialization skipped (serverless mode): {e}")

# Import analytics (always available)
from analytics import (
    analyze_trades, generate_report, compare_wallets,
    PRICING_TIERS, get_tier_info, check_feature_access, calculate_overage
)

# Constants
DATA_API = "https://data-api.polymarket.com"
OUTPUT_DIR = "/tmp/activity-exports"  # Vercel uses /tmp for writable storage
BATCH_SIZE = 1000
TRADE_PRESETS = [100, 1000, 5000, 10000, 50000, 100000]


def fetch_all_trades(wallet: str, callback=None) -> Dict:
    """Fetch ALL trades using backward timestamp pagination."""
    all_trades = []
    seen_keys: Set[tuple] = set()
    end_ts = None
    username = None
    batch_num = 0

    while True:
        batch_num += 1
        params = {"user": wallet, "limit": BATCH_SIZE}
        if end_ts:
            params["end"] = end_ts

        try:
            response = requests.get(f"{DATA_API}/activity", params=params, timeout=30)
            response.raise_for_status()
            trades = response.json()
        except Exception as e:
            break

        if not trades or not isinstance(trades, list):
            break

        if not username and trades[0].get('name'):
            username = trades[0].get('name')

        new_count = 0
        oldest_ts = float('inf')

        for t in trades:
            key = (t.get('transactionHash'), t.get('timestamp'), t.get('asset'))
            if key not in seen_keys:
                seen_keys.add(key)
                all_trades.append(t)
                new_count += 1

            ts = t.get('timestamp', 0)
            if ts and ts < oldest_ts:
                oldest_ts = ts

        if len(trades) < BATCH_SIZE or new_count == 0:
            break

        if oldest_ts != float('inf'):
            end_ts = oldest_ts
        else:
            break

        time.sleep(0.2)

        if batch_num > 5000:
            break

    all_trades.sort(key=lambda t: t.get('timestamp', 0), reverse=True)

    buys = sum(1 for t in all_trades if t.get('side') == 'BUY')
    sells = sum(1 for t in all_trades if t.get('side') == 'SELL')
    volume = sum(t.get('usdcSize', 0) or 0 for t in all_trades)

    timestamps = [t.get('timestamp', 0) for t in all_trades if t.get('timestamp')]
    oldest = datetime.fromtimestamp(min(timestamps)).strftime('%Y-%m-%d %H:%M') if timestamps else None
    newest = datetime.fromtimestamp(max(timestamps)).strftime('%Y-%m-%d %H:%M') if timestamps else None

    return {
        "wallet": wallet,
        "username": username,
        "fetched_at": datetime.now(timezone.utc).isoformat(),
        "trade_count": len(all_trades),
        "stats": {
            "buys": buys,
            "sells": sells,
            "volume_usd": volume,
            "oldest_trade": oldest,
            "newest_trade": newest,
            "batches": batch_num
        },
        "trades": all_trades
    }


def fetch_recent_trades(wallet: str, limit: int = 100) -> Dict:
    """Fetch recent trades (quick mode)."""
    params = {"user": wallet, "limit": limit}

    try:
        response = requests.get(f"{DATA_API}/activity", params=params, timeout=30)
        response.raise_for_status()
        trades = response.json() if isinstance(response.json(), list) else []

        username = None
        if trades and trades[0].get('name'):
            username = trades[0].get('name')

        return {
            "wallet": wallet,
            "username": username,
            "fetched_at": datetime.now(timezone.utc).isoformat(),
            "trade_count": len(trades),
            "trades": trades[:limit]
        }
    except Exception as e:
        return {"wallet": wallet, "error": str(e), "trades": [], "trade_count": 0}


def fetch_trades_with_limit(wallet: str, limit: int) -> Dict:
    """Fetch trades up to a specific limit using pagination."""
    all_trades = []
    seen_keys: Set[tuple] = set()
    end_ts = None
    username = None

    while len(all_trades) < limit:
        batch_limit = min(BATCH_SIZE, limit - len(all_trades))
        params = {"user": wallet, "limit": batch_limit}
        if end_ts:
            params["end"] = end_ts

        try:
            response = requests.get(f"{DATA_API}/activity", params=params, timeout=30)
            response.raise_for_status()
            trades = response.json()
        except Exception as e:
            break

        if not trades or not isinstance(trades, list):
            break

        if not username and trades[0].get('name'):
            username = trades[0].get('name')

        oldest_ts = float('inf')
        new_count = 0

        for t in trades:
            if len(all_trades) >= limit:
                break
            key = (t.get('transactionHash'), t.get('timestamp'), t.get('asset'))
            if key not in seen_keys:
                seen_keys.add(key)
                all_trades.append(t)
                new_count += 1

            ts = t.get('timestamp', 0)
            if ts and ts < oldest_ts:
                oldest_ts = ts

        if len(trades) < batch_limit or new_count == 0:
            break

        if oldest_ts != float('inf'):
            end_ts = oldest_ts
        else:
            break

        time.sleep(0.1)

    all_trades.sort(key=lambda t: t.get('timestamp', 0), reverse=True)

    return {
        "wallet": wallet,
        "username": username,
        "fetched_at": datetime.now(timezone.utc).isoformat(),
        "trade_count": len(all_trades),
        "trades": all_trades
    }


# Routes
@app.route('/')
def index():
    return send_from_directory('..', 'trade-viewer.html')


@app.route('/api/trades/<wallet>')
def get_trades(wallet):
    """Fetch trades - use mode=full for all history."""
    mode = request.args.get('mode', 'recent')
    limit = request.args.get('limit', 100, type=int)

    if mode == 'full':
        result = fetch_all_trades(wallet)
    elif limit > 1000:
        # Use paginated fetch for larger limits (5k, 10k, 50k, 100k)
        result = fetch_trades_with_limit(wallet, limit)
    else:
        result = fetch_recent_trades(wallet, limit)

    return jsonify(result)


@app.route('/api/download/<wallet>/<format>')
def download_trades(wallet, format):
    """Download trades as JSON, CSV, or TXT."""
    mode = request.args.get('mode', 'recent')
    limit = request.args.get('limit', 100, type=int)

    if mode == 'full':
        result = fetch_all_trades(wallet)
    elif limit > 1000:
        # Use paginated fetch for larger limits
        result = fetch_trades_with_limit(wallet, limit)
    else:
        result = fetch_recent_trades(wallet, limit)

    short_wallet = wallet[:10]
    username = result.get('username') or short_wallet
    suffix = '-full' if mode == 'full' else ''

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if format == 'json':
        filename = f"{OUTPUT_DIR}/{username}{suffix}-trades.json"
        with open(filename, 'w') as f:
            json.dump(result, f, indent=2)
        return send_file(filename, as_attachment=True, download_name=f"{username}{suffix}-trades.json")

    elif format == 'csv':
        output = io.StringIO()
        if result['trades']:
            fieldnames = ['timestamp', 'datetime', 'type', 'side', 'price', 'size',
                         'usdcSize', 'title', 'outcome', 'transactionHash']
            writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()

            for trade in result['trades']:
                ts = trade.get('timestamp', 0)
                trade['datetime'] = datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S') if ts else ''
                writer.writerow(trade)

        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode()),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f"{username}{suffix}-trades.csv"
        )

    elif format == 'txt':
        # Generate formatted text file
        lines = []
        lines.append("=" * 60)
        lines.append("POLYMARKET TRADE HISTORY")
        lines.append("=" * 60)
        lines.append("")
        lines.append(f"Wallet: {wallet}")
        lines.append(f"Username: {username}")
        lines.append(f"Total Trades: {len(result.get('trades', []))}")
        lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("")
        lines.append("-" * 60)
        lines.append("")

        for i, trade in enumerate(result.get('trades', []), 1):
            ts = trade.get('timestamp', 0)
            dt = datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S') if ts else 'N/A'
            side = trade.get('side', 'N/A')
            price = trade.get('price', 0)
            size = trade.get('size', 0)
            usdc = trade.get('usdcSize', 0)
            title = trade.get('title', 'N/A')
            outcome = trade.get('outcome', 'N/A')
            tx_hash = trade.get('transactionHash', 'N/A')

            lines.append(f"Trade #{i}")
            lines.append(f"  Time:     {dt}")
            lines.append(f"  Side:     {side}")
            lines.append(f"  Market:   {title}")
            lines.append(f"  Outcome:  {outcome}")
            lines.append(f"  Price:    {price:.4f}" if isinstance(price, (int, float)) else f"  Price:    {price}")
            lines.append(f"  Size:     {size:.2f}" if isinstance(size, (int, float)) else f"  Size:     {size}")
            lines.append(f"  Value:    ${usdc:.2f}" if isinstance(usdc, (int, float)) else f"  Value:    ${usdc}")
            lines.append(f"  Tx:       {tx_hash[:20]}..." if tx_hash and len(str(tx_hash)) > 20 else f"  Tx:       {tx_hash}")
            lines.append("")

        txt_content = "\n".join(lines)

        return send_file(
            io.BytesIO(txt_content.encode()),
            mimetype='text/plain',
            as_attachment=True,
            download_name=f"{username}{suffix}-trades.txt"
        )

    elif format == 'xlsx':
        # Generate Excel file using openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return jsonify({"error": "Excel export not available. Install openpyxl."}), 500

        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = ['#', 'Date/Time', 'Side', 'Market', 'Outcome', 'Price', 'Size', 'Value (USDC)', 'Transaction Hash']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        for row_num, trade in enumerate(result.get('trades', []), 2):
            ts = trade.get('timestamp', 0)
            dt = datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S') if ts else 'N/A'
            
            row_data = [
                row_num - 1,
                dt,
                trade.get('side', 'N/A'),
                trade.get('title', 'N/A'),
                trade.get('outcome', 'N/A'),
                trade.get('price', 0),
                trade.get('size', 0),
                trade.get('usdcSize', 0),
                trade.get('transactionHash', 'N/A')
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                if col in [6, 7, 8]:  # Price, Size, Value columns
                    cell.number_format = '#,##0.0000' if col == 6 else '#,##0.00'

        # Adjust column widths
        column_widths = [6, 20, 8, 50, 15, 12, 12, 15, 45]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Save to BytesIO
        excel_output = io.BytesIO()
        wb.save(excel_output)
        excel_output.seek(0)

        return send_file(
            excel_output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{username}{suffix}-trades.xlsx"
        )

    return jsonify({"error": "Invalid format. Use 'json', 'csv', 'txt', or 'xlsx'"}), 400


@app.route('/api/analyze/<wallet>')
def analyze_wallet(wallet):
    """Analyze trades and return pattern insights."""
    limit = request.args.get('limit', 1000, type=int)
    mode = request.args.get('mode', 'recent')

    if mode == 'full':
        result = fetch_all_trades(wallet)
    else:
        result = fetch_trades_with_limit(wallet, limit)

    trades = result.get('trades', [])
    if not trades:
        return jsonify({"error": "No trades found", "wallet": wallet})

    analysis = analyze_trades(trades)

    return jsonify({
        "wallet": wallet,
        "username": result.get('username'),
        "trade_count": len(trades),
        "analysis": analysis
    })


@app.route('/api/analyze/<wallet>/report')
def get_analysis_report(wallet):
    """Generate and download markdown analysis report."""
    limit = request.args.get('limit', 1000, type=int)
    mode = request.args.get('mode', 'recent')

    if mode == 'full':
        result = fetch_all_trades(wallet)
    else:
        result = fetch_trades_with_limit(wallet, limit)

    trades = result.get('trades', [])
    if not trades:
        return jsonify({"error": "No trades found"}), 404

    analysis = analyze_trades(trades)
    report = generate_report(analysis, wallet, result.get('username'))

    username = result.get('username') or wallet[:10]
    return Response(
        report,
        mimetype='text/markdown',
        headers={'Content-Disposition': f'attachment; filename={username}-analysis.md'}
    )


@app.route('/api/compare', methods=['POST'])
def compare_wallets_endpoint():
    """Compare multiple wallets side-by-side."""
    data = request.get_json()
    wallets = data.get('wallets', [])
    limit = data.get('limit', 1000)
    tier = data.get('tier', 'free')

    if not wallets:
        return jsonify({"error": "No wallets provided"}), 400

    tier_info = get_tier_info(tier)
    max_wallets = tier_info.get("comparison_wallets", 0)

    if len(wallets) > max_wallets and tier != 'scale':
        return jsonify({
            "error": f"Your plan allows comparing up to {max_wallets} wallets. Upgrade to compare more.",
            "tier": tier,
            "max_wallets": max_wallets,
            "upgrade_needed": True
        }), 403

    wallet_analyses = []
    for wallet in wallets[:max_wallets or len(wallets)]:
        result = fetch_trades_with_limit(wallet, limit)
        trades = result.get('trades', [])

        if trades:
            analysis = analyze_trades(trades)
            wallet_analyses.append({
                "wallet": wallet,
                "username": result.get('username'),
                "trade_count": len(trades),
                "analysis": analysis,
                "pnl": analysis.get("pnl", {})
            })

    comparison = compare_wallets(wallet_analyses)

    return jsonify({
        "comparison": comparison,
        "wallet_data": wallet_analyses,
        "tier": tier
    })


@app.route('/api/presets')
def get_presets():
    """Return available trade count presets."""
    return jsonify({"presets": TRADE_PRESETS})


@app.route('/api/pricing')
def get_pricing():
    """Return pricing tier information."""
    return jsonify({
        "tiers": PRICING_TIERS,
        "overage_rate": 0.01,
        "currency": "USD"
    })


@app.route('/api/pricing/<tier>')
def get_tier(tier):
    """Get specific tier information."""
    tier_info = get_tier_info(tier)
    if tier_info:
        return jsonify(tier_info)
    return jsonify({"error": "Invalid tier"}), 404


@app.route('/api/health')
def health_check():
    """Health check endpoint."""
    return jsonify({"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()})


# Error handlers
@app.errorhandler(404)
def not_found(e):
    return jsonify({"error": "Not found"}), 404


@app.errorhandler(500)
def server_error(e):
    return jsonify({"error": "Internal server error"}), 500


# For local development
if __name__ == '__main__':
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print("\n🚀 PolyEdge.io")
    print("   http://localhost:8080")
    app.run(debug=True, port=8080, host='0.0.0.0')
